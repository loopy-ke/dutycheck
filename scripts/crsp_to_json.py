"""
Convert KRA CRSP Excel (July 2025) to structured JSON files.
Run: python3 scripts/crsp_to_json.py

Outputs:
  data/crsp_vehicles.json      — all 5,200+ motor vehicle entries
  data/crsp_motorcycles.json   — motorcycle entries
  data/depreciation.json       — depreciation tables from TEMPLATE sheet
"""

import json
import openpyxl
from pathlib import Path

EXCEL_FILE = Path(__file__).parent.parent / "New-CRSP---July-2025.xlsx"
DATA_DIR = Path(__file__).parent.parent / "data"
DATA_DIR.mkdir(exist_ok=True)

wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)


# ── Motor Vehicles ──────────────────────────────────────────────────────────

def parse_vehicles():
    ws = wb["M.Vehicle CRSP July 2025"]
    vehicles = []
    headers_found = False

    for row in ws.iter_rows(values_only=True):
        # Skip until we find the header row
        if not headers_found:
            if row[0] == "Make":
                headers_found = True
            continue

        make, model, model_number, transmission, drive, engine_cc, body_type, gvw, seating, fuel, crsp, *_ = row

        # Skip empty rows
        if not make or not model or not crsp:
            continue

        vehicles.append({
            "make": str(make).strip(),
            "model": str(model).strip(),
            "model_number": str(model_number).strip() if model_number else None,
            "transmission": str(transmission).strip() if transmission else None,
            "drive": str(drive).strip() if drive else None,
            "engine_cc": engine_cc,  # int (cc) or string (e.g. "63 kWh" for EVs)
            "body_type": str(body_type).strip() if body_type else None,
            "gvw": gvw,
            "seating": seating,
            "fuel": str(fuel).strip() if fuel else None,
            "crsp_kes": round(float(crsp)),
        })

    print(f"Motor vehicles: {len(vehicles)} entries")
    return vehicles


# ── Motorcycles ─────────────────────────────────────────────────────────────

def parse_motorcycles():
    ws = wb["Motor Cycles July 2025"]
    bikes = []
    headers_found = False

    for row in ws.iter_rows(values_only=True):
        if not headers_found:
            if row[0] == "Make":
                headers_found = True
            continue

        make, model, model_number, transmission, engine_cc, seating, fuel, crsp = row[:8]

        if not make or not model or not crsp:
            continue

        bikes.append({
            "make": str(make).strip(),
            "model": str(model).strip(),
            "model_number": str(model_number).strip() if model_number else None,
            "transmission": str(transmission).strip() if transmission else None,
            "engine_cc": engine_cc,
            "seating": seating,
            "fuel": str(fuel).strip() if fuel else None,
            "crsp_kes": round(float(crsp)),
        })

    print(f"Motorcycles: {len(bikes)} entries")
    return bikes


# ── Depreciation Tables ─────────────────────────────────────────────────────
# From TEMPLATE 2025 sheet rows 3-11 (0-indexed: rows 2-10)

def parse_depreciation():
    ws = wb["TEMPLATE 2025"]
    rows = list(ws.iter_rows(values_only=True))

    # Direct imports depreciation (columns B-C, rows 3-11)
    direct = []
    previously_registered = []

    for row in rows[2:]:  # skip 2 header rows
        label_direct = row[1]
        pct_direct = row[2]
        label_prev = row[8]
        pct_prev = row[9]

        if label_direct and pct_direct is not None:
            direct.append({"label": str(label_direct).strip(), "rate": float(pct_direct)})
        if label_prev and pct_prev is not None:
            previously_registered.append({"label": str(label_prev).strip(), "rate": float(pct_prev)})

        # Stop when both run out
        if not label_direct and not label_prev:
            break

    return {
        "direct_imports": direct,
        "previously_registered": previously_registered,
    }


# ── Duty Rates (hardcoded from KRA, these don't change with CRSP) ────────────

DUTY_RATES = {
    "import_duty_rate": 0.25,       # 25% of customs value
    "excise_duty_rate": 0.20,       # 20% of (CV + import duty)
    "vat_rate": 0.16,               # 16% of (CV + import duty + excise)
    "idf_rate": 0.0225,             # 2.25% of CV (min KES 5,000)
    "rdl_rate": 0.015,              # 1.5% of CV
    "idf_minimum_kes": 5000,
    "crsp_tax_strip_divisor": 2.4469,  # strips built-in taxes from CRSP to get customs value
    "note": "Rates as per KRA Finance Act 2025. Tax-strip divisor derived from official KRA valuation template."
}


# ── Write outputs ────────────────────────────────────────────────────────────

if __name__ == "__main__":
    vehicles = parse_vehicles()
    motorcycles = parse_motorcycles()
    depreciation = parse_depreciation()

    (DATA_DIR / "crsp_vehicles.json").write_text(
        json.dumps(vehicles, indent=2, ensure_ascii=False)
    )
    print(f"Written → data/crsp_vehicles.json")

    (DATA_DIR / "crsp_motorcycles.json").write_text(
        json.dumps(motorcycles, indent=2, ensure_ascii=False)
    )
    print(f"Written → data/crsp_motorcycles.json")

    (DATA_DIR / "depreciation.json").write_text(
        json.dumps(depreciation, indent=2, ensure_ascii=False)
    )
    print(f"Written → data/depreciation.json")

    (DATA_DIR / "duty_rates.json").write_text(
        json.dumps(DUTY_RATES, indent=2, ensure_ascii=False)
    )
    print(f"Written → data/duty_rates.json")

    print("\nDone. All CRSP data extracted to /data/")
